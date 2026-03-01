"""Read CoastSnap site metadata from the Excel database.

The CoastSnap MATLAB toolbox stores one monitoring *site* per sheet in
``CoastSnapDB.xlsx``. Each sheet contains station/rectification
settings plus a "Ground Control Points" section:

    Ground Control Points
    GCP name, <name>
    Eastings, <value> [units...]
    Northings, <value> [units...]
    Elevation, <value> [units...]

Values often include **comma decimals** and **units/text** (e.g.
``341761,07 m MGA94``). This module parses those robustly and also
parses the MATLAB-style ``GCP combo`` (e.g. ``[1:7]`` or ``[1 2 5:9]``).

We prefer :mod:`pandas` to read the workbook, but fall back to
:mod:`openpyxl` if pandas isn't installed.
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple
from pathlib import Path
import json
import re

try:
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None  # type: ignore

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore


IGNORE_SHEETS = {
    "database",
    "tidal offset calculation",
    "insert_newsitename_here",
}


def _to_float(val: Any) -> Optional[float]:
    """Parse numbers like '341761,07 m MGA94' -> 341761.07."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    m = re.search(r"[-+]?\d*\.?\d+", s)
    if not m:
        return None
    try:
        return float(m.group())
    except Exception:
        return None


def _find_label_in_row(row: List[Any], label: str, max_cols: int = 8) -> Tuple[Optional[str], Optional[int]]:
    lab = label.strip().lower()
    for c in range(min(max_cols, len(row))):
        v = row[c]
        if isinstance(v, str) and v.strip().lower() == lab:
            return v, c
    return None, None


def _parse_matlab_index_list(text: str) -> List[int]:
    s = (text or "").strip().strip("[]")
    if not s:
        return []
    s = s.replace(",", " ")
    parts = s.split()
    out: List[int] = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            try:
                a, b = part.split(":", 1)
                start = int(a)
                end = int(b)
            except Exception:
                continue
            step = 1 if start <= end else -1
            out.extend(list(range(start, end + step, step)))
        else:
            try:
                out.append(int(part))
            except Exception:
                continue
    # de-duplicate preserving order
    seen = set()
    ordered: List[int] = []
    for i in out:
        if i not in seen:
            ordered.append(i)
            seen.add(i)
    return ordered


def _parse_sheet_rows(sheet_name: str, rows: List[List[Any]]) -> Dict[str, Any]:
    meta: Dict[str, Any] = {"site_name": sheet_name}

    # Generic key/value pairs often sit in col 0/1
    for r in rows:
        if not r:
            continue
        key = r[0]
        if isinstance(key, str) and key.strip() and len(r) > 1 and r[1] not in (None, ""):
            k = key.strip()
            if k.lower() in {"ground control points", "gcp rectification combo"}:
                continue
            if k.lower() == "gcp name":
                continue
            meta[k] = r[1]

    # Find "Ground Control Points" header
    gcp_start = None
    for i, r in enumerate(rows):
        for c in range(min(8, len(r))):
            v = r[c]
            if isinstance(v, str) and v.strip().lower() == "ground control points":
                gcp_start = i
                break
        if gcp_start is not None:
            break

    gcp_world: List[Dict[str, Any]] = []
    if gcp_start is not None:
        i = gcp_start + 1
        while i < len(rows):
            r0 = rows[i]
            _, c0 = _find_label_in_row(r0, "gcp name")
            if c0 is None:
                break
            name_val = r0[c0 + 1] if c0 + 1 < len(r0) else None

            r_e = rows[i + 1] if i + 1 < len(rows) else []
            r_n = rows[i + 2] if i + 2 < len(rows) else []
            r_z = rows[i + 3] if i + 3 < len(rows) else []

            _, ce = _find_label_in_row(r_e, "eastings")
            _, cn = _find_label_in_row(r_n, "northings")
            _, cz = _find_label_in_row(r_z, "elevation")

            east_val = r_e[ce + 1] if ce is not None and ce + 1 < len(r_e) else None
            north_val = r_n[cn + 1] if cn is not None and cn + 1 < len(r_n) else None
            elev_val = r_z[cz + 1] if cz is not None and cz + 1 < len(r_z) else None

            east = _to_float(east_val)
            north = _to_float(north_val)
            elev = _to_float(elev_val)

            if name_val is not None and east is not None and north is not None:
                gcp_world.append(
                    {
                        "name": str(name_val).strip(),
                        "eastings": float(east),
                        "northings": float(north),
                        "elevation": float(elev) if elev is not None else 0.0,
                    }
                )
                i += 4
            else:
                i += 1

    if gcp_world:
        meta["gcp_world"] = gcp_world

    # Find GCP combo anywhere in sheet
    for r in rows:
        _, c = _find_label_in_row(r, "gcp combo")
        if c is None:
            continue
        combo_val = r[c + 1] if c + 1 < len(r) else None
        combo = _parse_matlab_index_list(str(combo_val) if combo_val is not None else "")
        if combo:
            meta["gcp_combo"] = combo
        break

    return _postprocess_meta(meta)


def _postprocess_meta(meta: Dict[str, Any]) -> Dict[str, Any]:
    # Convert numeric strings with units to floats where safe
    for k, v in list(meta.items()):
        if k in {"gcp_world", "gcp_combo", "site_name"}:
            continue
        if isinstance(v, str):
            num = _to_float(v)
            if num is not None:
                meta[k] = num
    return meta


def read_site_database(path: str) -> Dict[str, Any]:
    """Read a CoastSnap site database (Excel or JSON)."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Site database file not found: {path}")

    ext = p.suffix.lower()
    if ext in {".xlsx", ".xls"}:
        sites: Dict[str, Any] = {}

        if pd is not None:
            xls = pd.ExcelFile(str(p))
            for sheet in xls.sheet_names:
                sheet_key = str(sheet).strip().lower()
                if any(ign in sheet_key for ign in IGNORE_SHEETS):
                    continue
                df = xls.parse(sheet, header=None)
                rows = df.values.tolist()
                sites[sheet_key] = _parse_sheet_rows(str(sheet).strip(), rows)
            return sites

        if openpyxl is None:
            raise ImportError(
                "Unable to read CoastSnapDB.xlsx. Install 'pandas' (recommended) or 'openpyxl'."
            )
        wb = openpyxl.load_workbook(str(p), data_only=True)
        for sheet in wb.sheetnames:
            sheet_key = str(sheet).strip().lower()
            if any(ign in sheet_key for ign in IGNORE_SHEETS):
                continue
            ws = wb[sheet]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            sites[sheet_key] = _parse_sheet_rows(str(sheet).strip(), rows)
        return sites

    if ext == ".json":
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return {str(k).strip().lower(): v for k, v in data.items()}
        if isinstance(data, list):
            out: Dict[str, Any] = {}
            for entry in data:
                if not isinstance(entry, dict) or "site_id" not in entry:
                    continue
                out[str(entry["site_id"]).strip().lower()] = entry
            return out
        raise ValueError("Site database JSON must be a dictionary or a list")

    raise ValueError(f"Unsupported database format: {ext}")


__all__ = ["read_site_database"]
